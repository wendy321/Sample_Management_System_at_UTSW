#!/usr/bin/python

from mysql.connector import Error
from openpyxl import load_workbook
from utility import *

""" Get XLSX Blob from RemoteR database """
def getXLSXBlobInDB(job_id, tag, data_type):
	return read_excel_blob_from_remoter_db(job_id, tag, data_type)

tmp_xlsx_filename = 'temp_excel.xlsx'
sample_ws_name = 'sample_batch_upload_example'
try:
    print('Read Excel blob from remote r db')
    excel = getXLSXBlobInDB(job_id = "test",tag = '1',data_type = 'sample')
    print('Write Excel blob to file in the disk')
    write_file(excel, tmp_xlsx_filename)
except Error as error:
    print(error) 
finally:
    print('Finish writing excel file')

wb = load_workbook(tmp_xlsx_filename)
print wb.sheetnames

