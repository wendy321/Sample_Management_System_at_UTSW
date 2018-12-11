#!/usr/bin/python

# from unicodedata import *
from mysql.connector import Error
from openpyxl import load_workbook
#from utility import *
import os, sys, uuid
from utility import write_file
from utility import read_excel_blob_from_remoter_db, \
		    read_last_sampleid_from_sms_db, \
		    read_uuid_from_sms_db_by_local_sample_id, \
		    read_uuid_from_sms_db_by_study_sample_id, \
			read_sample_from_sms_db_by_local_sample_id, \
					read_sample_from_sms_db_by_study_sample_id, \
					read_patient_id_from_sms_db_by_local_patient_id, \
					read_patient_id_from_sms_db_by_study_patient_id, \
					read_patient_from_sms_db_by_local_patient_id, \
					read_patient_from_sms_db_by_study_patient_id
from utility import insert_sample_into_sms_db, \
					insert_patient_into_sms_db, \
					insert_enrollstudy_into_sms_db, \
					insert_success_result_into_remoteR_db, \
					insert_error_result_into_remoteR_db, \
					insert_warn_result_into_remoteR_db
from utility import update_local_sample_id_by_uuid_in_sms_db, \
					update_local_patient_id_by_patid_in_sms_db, \
					update_enrollstudy_into_sms_db
from utility import generatePatientID, generateSampleID


""" Global Variable """
micorUnic = u"\u03BC" # micorUTF8 = "\xce\xbc"
micor = micorUnic.encode('utf-8', errors='replace')
eUnic =u"\u00E9" # eUTF8 = "\xc3\xa9"
e = eUnic.encode('utf-8', errors='replace')
codeDic = {
	0: {'AGCT0132': 0, 'AGCT01P1': 4, 'AGCT0521': 14, 'GC 1': 9, 'GC 2': 8, 'GOG 0078': 5, 'GOG 0090': 1, 'GOG 0116': 11,'INT-0097': 7,
				'INT-0106': 10,'OPTF': 3,'P9749': 2,'TCG_99': 13,'TGM95': 12,'Other': 98,'Unknown': 99,'Not In Clinical Trial':100, 'None': None},
	6: {'DNA': 1, 'RNA': 2, 'Protein': 3, 'Tissue':	4, 'Cell': 5, 'Fluid': 6, 'Other': 98, 'Unknown':99, 'None': None},
	7: {
		'DNA, Whole Genome Amplified DNA': 11,
		'DNA, cDNA': 12,
		'DNA, ctDNA': 13,
		'DNA, Genomic DNA': 14,
		'DNA, Not Specified': 19,
		'RNA, poly-A enriched': 21,
		'RNA, Nuclear': 22,
		'RNA, Cytoplasmic': 23,
		'RNA, Total RNA': 24,
		'RNA, Not Specified': 29,
		'Protein, Not Specified': 39,
		'Tissue, Tissue Block': 41,
		'Tissue, Tissue Slide': 42,
		'Tissue, Microdissected': 43,
		'Tissue, Not Specified': 49,
		'Cell, Pleural Effusion All Cells': 51,
		'Cell, Pleural Effusion White Blood Cells': 52,
		'Cell, Peripheral Blood All Cells': 53,
		'Cell, Peripheral Blood White Cells': 54,
		'Cell, Peripheral Blood Mononuclear Cell (PBMC)': 55,
		'Cell, Cell Pellet': 56,
		'Cell, Not Specified':  59,
		'Fluid, Whole Blood': 61,
		'Fluid, Plasma': 62,
		'Fluid, Serum': 63,
		'Fluid, Bone Marrow': 64,
		'Fluid, Urine': 65,
		'Fluid, Saliva': 66,
		'Fluid, Cerebrospinal Fluid': 67,
		'Fluid, Pleural Fluid': 68,
		'Fluid, Ascites': 69,
		'Fluid, Lavage': 610,
		'Fluid, Body Cavity Fluid': 611,
		'Fluid, Milk': 612,
		'Fluid, Vitreous Fluid': 613,
		'Fluid, Gastric Fluid': 614,
		'Fluid, Amniotic Fluid': 615,
		'Fluid, Bile': 616,
		'Fluid, Synovial Fluid': 617,
		'Fluid, Sweat': 618,
		'Fluid, Feces':	619,
		'Fluid, Buffy Coat': 620,
		'Fluid, Sputum': 621,
		'Fluid, Not Specified': 699,
		'Other': 98,
		'Unknown': 99,
		'None': None
	},
	8: {
		'Primary Solid Tumor': 1,
		'Recurrent Solid Tumor': 2,
		'Metastatic Tumor':	3,
		'Blood Derived': 10,
		'Solid Tissue Normal': 11,
		'Other': 98,
		'Unknown': 99,
		'None': None
	},
	9: {
		'Amatruda Lab': 1,
		'Cold Room': 2,
		'Unknown': 99,
		'None': None
	},
	10: {
		'Freezer': 1,
		'Storage Cabinet': 2,
		'Unknown': 99,
		'None': None
	},
	11: {
		'-20 Celcius Degree': 1,
		'-80 Celcius Degree': 2,
		'Unknown': 99,
		'None': None
	},
	18: {
		micor+"g": 1,
		'mg': 2,
		'g': 3,
		micor+"L": 4,
		'mL': 5,
		'scrolls': 6,
		'cassettes': 7,
		'slides': 8,
		'blocks': 9,
		'unspecified': 10,
		'None': None
	},
	21: {
		'Flash Frozen': 1,
		'Frozen with OCT': 2,
		'FFPE': 3,
		'Fresh': 4,
		'Other': 98,
		'Unknown': 99,
		'None': None
	},
	22: {
		'CNS,Central Nervous System': '1',
		'CNS,Suprasellar\/Neurohypophyseal': '1.a',
		'CNS,Pineal': '1.b',
		'CNS,Bifocal': '1.c',
		'CNS,Thalamic': '1.d',
		'CNS,Cerebral Cortex': '1.e',
		'CNS,Cerebral Cortex,Frontal': '1.e.a',
		'CNS,Cerebral Cortex,Temporal': '1.e.b',
		'CNS,Cerebral Cortex,Parietal':	'1.e.c',
		'CNS,Cerebral Cortex,Occipital': '1.e.d',
		'CNS,Cerebral Cortex,Other': '1.e.y',
		'CNS,Spinal': '1.f',
		'CNS,Other': '1.y',
		'Head, neck (not CNS)': '2',
		'Liver (primary not metastatic)': '3',
		'Mediastinum': '4',
		'Ovary': '5',
		'Retroperitoneum': '6',
		'Sacrococcygeal': '7',
		'Testis': '8',
		'Vagina (female only)': '9',
		'Other': '98',
		'Unknown': '99',
		'None': None
	},
	23: {
		'Left': 1,
		'Right': 2,
		'Bilateral': 3,
		'Unknown': 99,
		'None': None
	},
	24: {
		'> 200bp': 1,
		'<= 200bp': 2,
		'Unknown': 99,
		'Not Applicable': 100,
		'None': None
	},
	25: {
		'Biopsy': 1,
		'Surgery': 2,
		'Blood collection': 3,
		'Saliva collection': 4,
		'Skin biopsy': 5,
		'Lumbar puncture': 6,
		'Other': 98,
		'Unknown': 99,
		'None': None
	},
	27: {
		'COG, MaGIC': 1,
		'CCLG, MaGIC': 2,
		'MRC': 3,
		'NRG Oncology': 4,
		'Other': 98,
		'Unknown': 99,
		'None': None
	},
	28: {
		'UT Southwestern Medical Center': 1,
		'Biopathology Center': 2,
		'Boston Children\'s Hospital': 3,
		'Children\'s Medical Center at Dallas':	4,
		'Cooperative Human Tissue Network':	5,
		'Erasmus Medical Center': 6,
		'Indiana University': 7,
		'Sant Joan de D'+e+'u Barcelona Children\'s Hospital': 8,
		'Other': 98,
		'Unknown': 99,
		'None': None
	}
}

sampleVariableDic = {
	0: 'Study_ID',
	1: 'Within_Study_Sample_ID',
	2: 'Within_Study_Patient_ID',
	3: 'Local_Sample_ID',
	4: 'Local_Patient_ID',
	5: 'Patient_ID',
	6: 'Sample_Class',
	7: 'Sample_Type',
	8: 'Pathological_Status',
	9: 'Storage_Room',
	10: 'Cabinet_Type',
	11: 'Cabinet_Temperature',
	12: 'Cabinet_Number',
	13: 'Shelf_Number',
	14: 'Rack_Number',
	15: 'Box_Number',
	16: 'Position_Number',
	17: 'Quantity_Value',
	18: 'Quantity_Unit',
	19: 'Concentration_Value',
	20: 'Concentration_Unit',
	21: 'Specimen_Type',
	22: 'Anatomical_Site',
	23: 'Anatomical_Laterality',
	24: 'Nucleotide_Size_Group_200',
	25: 'Procedure_Type',
	26: 'Date_Procedure',
	27: 'Sample_Contributor_Consortium_ID',
	28: 'Sample_Contributor_Institute_ID',
	29: 'Date_Derive_From_Parent',
	30: 'Notes'
}

""" Get XLSX Blob from RemoteR database """
def getXLSXBlobInDB(job_id, tag, data_type):
	return read_excel_blob_from_remoter_db(job_id, tag, data_type)

""" Validate data format """
def isValidDataFormat(colindx, inputstr):
	isValid = True;
	# switch(colindx){
	# 	case 0: case 5: case 6: ...: isValid=transferStrToCode(inputstr); break;
	# 	case 1: case : case : ...: isValid=checkStrLen(inputstr, maxlen); break;
	# 	case : case : case : ...: isValid=checkFloatLen(inputstr, maxlenWhole, maxlenDeci); break;
	# 	case : case : case : ...: isValid=checkDateFormat(inputstr); break;
	# 	default:
	# }
	return isValid;

""" Convert data to code """
def convertStrToCode(colindx, inputstr):
	return codeDic.get(colindx).get(inputstr)

""" Generate 36-character uuid """
def generate36CharUUID():
	return str(uuid.uuid4())

""" Get the sample UUID exists in Sample table in database """	
def getSampleUUIDInDBbyLocalSamID(local_sample_id):
	return read_uuid_from_sms_db_by_local_sample_id(local_sample_id)

""" Get the sample UUID in EnrollStudy table in database """
def getSampleUUIDInDBbyStudy(study_id,study_sample_id):
	return read_uuid_from_sms_db_by_study_sample_id(study_id,study_sample_id)

""" Get the sample row in Sample table in database """
def getSampleRowInDBbyLocalSamID(local_sample_id):
	return read_sample_from_sms_db_by_local_sample_id(local_sample_id)

""" Get the sample row in EnrollStudy table in database """
def getSampleRowInDBbyStudy(study_id,study_sample_id):
	return read_sample_from_sms_db_by_study_sample_id(study_id,study_sample_id)

""" Add a sample row in Sample table in database """
def addSampleInDB(one_sample_data):
	insert_sample_into_sms_db(one_sample_data)

""" Get the patient id in Patient table in database """
def getPatientIDInDBbyLocalPatID(local_patient_id):
	return read_patient_id_from_sms_db_by_local_patient_id(local_patient_id)

""" Get the patient id in EnrollStudy table in database """
def getPatientIDInDBbyStudy(study_id,study_patient_id):
	return read_patient_id_from_sms_db_by_study_patient_id(study_id,study_patient_id)

""" Get the patient row in Patient table in database """
def getPatientRowInDBbyLocalPatID(local_patient_id):
	return read_patient_from_sms_db_by_local_patient_id(local_patient_id)

""" Get the patient row in EnrollStudy table in database """
def getPatientRowInDBbyStudy(study_id,study_patient_id):
	return read_patient_from_sms_db_by_study_patient_id(study_id,study_patient_id)

""" Add a patient row with patient_id and local_patient_id in Patient table in database """
def addPatientInDB(patient_id, local_patient_id):
	insert_patient_into_sms_db(patient_id, local_patient_id)

""" Update the local_patient_id in Patient table in database """
def updateLocalPatientIDInDB(local_patient_id, patient_id):
	return update_local_patient_id_by_patid_in_sms_db(local_patient_id, patient_id)

""" Add an enroll study row with enroll study data in EnrollStudy table in database """
def addEnrollStudyInDB(enroll_study_data):
	return insert_enrollstudy_into_sms_db(enroll_study_data)

""" Update enroll study row(s) with enroll study data in EnrollStudy table in database 
:param enroll_study_data: enroll study data
:param by_variable: "source_sample_id" or "source_patient_id" 
					update the row(s) according to "source_sample_id" or "source_patient_id" and "source_id"
"""
def updateEnrollStudyInDB(enroll_study_data, by_variable):
	return update_enrollstudy_into_sms_db(enroll_study_data, by_variable)

""" Add Success Result into RemoteR database (one row for each sample, if success) """
def addSucessResultIntoRemoteRDB(job_id, sample_uuid, is_new_patient, patient_id, is_new_enroll_study, enroll_study_id):
	insert_success_result_into_remoteR_db(job_id, sample_uuid, is_new_patient, patient_id, is_new_enroll_study, enroll_study_id)

""" Add Error Result into RemoteR database (one row for each batch , if error) """
def addErrorResultIntoRemoteRDB(job_id, msg):
	insert_error_result_into_remoteR_db(job_id, msg)

""" Add Warn Result into RemoteR database (one row for each batch , if warn) """
def	addWarnResultIntoRemoteRDB(job_id, msg, existed_sam_uuids):
	insert_warn_result_into_remoteR_db(job_id, msg, existed_sam_uuids)

""" Process patient related data
(Patient, and EnrollStudy tables are involved)
:param rowNum: row number in the xlsx file
:param sample_data: validate and converted sample data in the xlsx file
:return: result of processing patient related data. 
If has error_msg, abort the whole batch processing, and store the error_msg into RemoteR and SMS databases.
If no error_msg, contintue the process
"""
def processPatient(rowNum, sample_data):
	result = {"Error_Msg": None, "Patient_ID": None, "Pat_Ops": None, "EnrollStudy_Ops": None}

	isError = False
	errorMsg = ""
	
	system_patient_id = sample_data['Patient_ID']
	patient_id_Patient_Table = getPatientIDInDBbyLocalPatID(sample_data['Local_Patient_ID'])
	source_id = sample_data['Study_ID']
	source_patient_id = sample_data['Within_Study_Patient_ID']
	if(source_patient_id != None):
		patient_row = getPatientRowInDBbyStudy(source_id,source_patient_id)
		patient_id_EnrollStudy_Table = None
		if(patient_row != None):
			patient_id_EnrollStudy_Table = patient_row[3]

	# compare input system_patient_id with patient_id queried by local_patient_id from Patient table 
	if((patient_id_Patient_Table != None) and (system_patient_id != None) and (system_patient_id != patient_id_Patient_Table)):
		isError = True
		errorMsg = errorMsg.join("Row number: %i, conflict system_patient_ids, i.e. " %(rowNum)) \
									.join("the input system_patient_id is different from ") \
									.join("the system_patient_id queried from DB by input local_patient_id.") \
									.join(" Please modify system_patient_id or local_patient_id. \n")
	else:
		if(patient_id_Patient_Table != None):
			# compare previous patient_id with the patient_id queried by source_id and source_patient_id from EnrollStudy table
			if(source_patient_id != None):
				if(patient_id_EnrollStudy_Table != None):
					if(patient_id_Patient_Table != patient_id_EnrollStudy_Table):
						isError = True
						errorMsg = errorMsg.join("Row number: %i, conflict system_patient_ids, i.e." %(rowNum)) \
											.join("the system_patient_id queried from DB by input local_patient_id and ") \
											.join("that by input source_name and source_patient_id are different.") \
											.join(" Please modify local_patient_id or source_name, source_patient_id. \n")
					else:
						result["EnrollStudy_Ops"] = "no_action" 
				else:
					if(patient_row != None):
						# update Patient_ID, Sample_UUID w/i known source_id & known source_patient_id in EnrollStudy table later after new sample has generated
						result["EnrollStudy_Ops"] = "update"
					else:
						# insert a new enrollstudy record w/i Patient_ID, Sample_UUID, and knonw source_id & known source_patient_id after new sample has generated
						result["EnrollStudy_Ops"] = "insert"
			else:
				# insert a new enrollstudy record w/i Patient_ID, Sample_UUID, and knonw source_id after new sample has generated
				result["EnrollStudy_Ops"] = "insert"

			if(not isError):
				result["Patient_ID"] = patient_id_Patient_Table
				result["Pat_Ops"] = "no_action"
		else:
			if((system_patient_id != None) and (patient_id_EnrollStudy_Table != None) and (system_patient_id != patient_id_EnrollStudy_Table)):
				isError = True
				errorMsg = errorMsg.join("Row number: %i, conflict system_patient_ids, i.e." %(rowNum)) \
									.join("the input system_patient_id is different with ") \
									.join("the system_patient_id queried from DB by input source_id and source_patient_id.") \
									.join(" Please modify system_patient_id or source_name, source_patient_id. \n")
			else:
				if(patient_id_EnrollStudy_Table != None):
					result["Patient_ID"] = patient_id_EnrollStudy_Table
					result["Pat_Ops"] = "insert"
					result["EnrollStudy_Ops"] = "no_action"
				elif((system_patient_id != None) and (patient_id_EnrollStudy_Table == None)):
					result["Patient_ID"] = system_patient_id
					result["Pat_Ops"] = "update"
					result["EnrollStudy_Ops"] = "insert"
				else:
					result["Patient_ID"] = generatePatientID(is_unlinked_patient_id = True)
					result["Pat_Ops"] = "insert"
					result["EnrollStudy_Ops"] = "insert"

	if(isError):
		result["Error_Msg"] = errorMsg

	return result

""" Process sample related data in database
(Sample, Patient, and EnrollStudy tables are involved)
:param rowNum: row number in the xlsx file
:param sample_data: validate and converted sample data in the xlsx file
:return: result of processing sample related data.
If has error_msg, abort the whole batch processing, and store the error_msg into RemoteR and SMS databases.
If no error_msg, contintue the process.
If has warn_msg, continue to the next sample 
"""
def processSample(rowNum, sample_data):
	result = {"Error_Msg": None, "Warn_Msg": None, "Existed_Sam_UUID": None, "Sam_Ops": None, "EnrollStudy_Ops": None}

	isError = False

	source_id = sample_data.get('Study_ID')
	source_sample_id = sample_data.get('Within_Study_Sample_ID')
	local_sample_id = sample_data.get('Local_Sample_ID')
	if(source_sample_id != None):
		sample_row = getSampleRowInDBbyStudy(source_id,source_sample_id)
		if(sample_row != None and sample_row[5] != None):
			sample_uuid = sample_row[5]
			update_result = update_local_sample_id_by_uuid_in_sms_db(local_sample_id, sample_uuid)
			if(update_result != None):
				result["Existed_Sam_UUID"] = sample_uuid
				result["Warn_Msg"] = "Row number: %i, source sample id %s has already existed in database.\n" %(rowNum, sample_data.get('SourceSampleID'))
			else:
				isError=True
				result["Error_Msg"] = "Row number: %i, couldn't connet to database, please inform developer and try later.\n" %(rowNum)
		elif(sample_row != None and sample_row[5] == None):
			# Currently, this case is impossible to generate based on our web design logic. 
			# If it happened, generate a new sample record w/i UUID and known src_sam_id, 
			# then update UUID with knonw source_id and known source_sample_id in EnrollStudy table.
			# (pat_id -> new sam_uuid -> update old enrollstudy_id w/i known src_id & src_sam_id)
			result["Sam_Ops"] = "insert"
			result["EnrollStudy_Ops"] = "update"
		else: 
			# Generate a new sample record w/i UUID and known src_sam_id 
			# (pat_id -> new sam_uuid -> new enrollstudy_id w/i known src_id & known src_sam_id)
			result["Sam_Ops"] = "insert"
			result["EnrollStudy_Ops"] = "insert"
	else: 
		# Generate a new sample record w/i UUID and Null src_sam_id 
		# (pat_id -> new sam_uuid -> new enrollstudy_id w/i known src_id & Null src_sam_id)
		# p.s. (10/02/2018) If user don't know src_pat_id, he/she will input src_pat_id_Pat as src_pat_id
		result["Sam_Ops"] = "insert"
		result["EnrollStudy_Ops"] = "insert"

	return result

def main():
	job_id = sys.argv[1]
	###################################################
	# Read Sample Blob from RemoteR DB into XLSX file
	###################################################

	tmp_xlsx_filename = 'temp_excel.xlsx'
	sample_ws_name = 'sample_batch_upload_example'
	try:
		print('Read Excel blob from remote r db')
		excel = getXLSXBlobInDB(job_id = job_id,tag = '1',data_type = 'sample')
		print('Write Excel blob to file in the disk')
		write_file(excel, tmp_xlsx_filename)
	except Error as error:
		print(error) 
	finally:
		print('Finish writing excel file')

	###################################################
	# Read Sample Data from XLSX file into Memory
	###################################################
	wb = load_workbook(tmp_xlsx_filename)
	# print wb.sheetnames
	ws = wb[sample_ws_name]
	rowNum = 1
	existedUUIDs = ""
	warnMsg = ""
	isError = False
	errorMsg = ""
	# for row in ws.iter_rows(min_row=1, max_col=31, max_row=2):
	for row in ws.iter_rows(min_row=2):
		result_sam = None
		isError = False

		# A sample dictionary which stores the validated and converted cell values of each sample row
		sample_data = {}

		# required cells
		study_name = row[0].value
		local_sample_id = row[3].value
		local_patient_id = row[4].value

		if((study_name != None) and (local_sample_id != None) and (local_patient_id != None)):
			existed_sam_uuid = getSampleUUIDInDBbyLocalSamID(local_sample_id = local_sample_id)
			if(existed_sam_uuid != None):
				warnMsg = warnMsg.join("Row number: %i, local sample id %s has already existed in database.\n" %(rowNum, local_sample_id))
				if(existedUUIDs == ""):
					existedUUIDs += existed_sam_uuid
				else:
					existedUUIDs += "," + existed_sam_uuid
				continue

			colNum = 0
			for cell in row:
				unic = u"%s" %(cell.value)
				w = unic.encode('utf-8', errors='replace')
				if(isValidDataFormat(colNum, w)):
					if(colNum in codeDic):
						finalW=convertStrToCode(colNum,w)
					else:
						finalW=w

					if(finalW == "None"):
						finalW = None

					sample_data[sampleVariableDic[colNum]] = finalW
					colNum += 1
				else: 
					isError=True	
					errorMsg = errorMsg.join("Row number: %i, Col number: %i, wrong data format.\n" %(rowNum, colNum))
					break

			result_sam = processSample(rowNum, sample_data)
			warn_sam = result_sam["Warn_Msg"]
			if(warn_sam != None):
				warnMsg = warnMsg.join(warn_sam)
				if(existedUUIDs == ""):
					existedUUIDs += result_sam["Existed_Sam_UUID"]
				else:
					existedUUIDs += "," + result_sam["Existed_Sam_UUID"]
				continue

			error_sam = result_sam["Error_Msg"]
			if(error_sam != None):
				isError = True
				errorMsg = errorMsg.join(error_sam)
				break

		else:
			break

		############################################################
		# Update or Insert Patient, Sample, EnrollStudy data into DB
		############################################################
		# process Patient record
		"""
		initial result_pat = {"Error_Msg": None, "Patient_ID": None, "Pat_Ops": None, "EnrollStudy_Ops": None}
		"""
		result_pat = processPatient(rowNum, sample_data)
		error_pat = result_pat["Error_Msg"]
		if(error_pat != None):
			isError = True
			errorMsg = errorMsg.join(error_pat)
			break

		is_new_patient = False
		pat_ops = result_pat["Pat_Ops"]
		pat_id = result_pat["Patient_ID"]
		if(pat_ops == "insert"):
			is_new_patient = True
			addPatientInDB(pat_id, local_patient_id)

		if(pat_ops == "update"):
			is_success = updateLocalPatientIDInDB(local_patient_id, pat_id)
			if(not is_success):
				isError = True
				errorMsg = errorMsg.join("Please contact developer. Fail to update local_patient_id into database based on your input.")
				break

		# process Sample record
		"""
		initial result_sam = {"Error_Msg": None, "Warn_Msg": None, "Existed_Sam_UUID": None, "Sam_Ops": None, "EnrollStudy_Ops": None}
		"""
		sam_ops = result_sam["Sam_Ops"]
		if(sam_ops == "insert"):
			sample_data["UUID"] = generate36CharUUID()
			sample_data["Patient_ID"] = pat_id
			if(sample_data["Pathological_Status"] == None):
				sample_data["Pathological_Status"] = 99			
			if(sample_data["Sample_Class"] == None):
				sample_data["Sample_Class"] = 99
			sample_data["Sample_ID"] = generateSampleID(pat_id, sample_data["Pathological_Status"], sample_data["Sample_Class"])
			addSampleInDB(sample_data) 

		# process Enroll Study record
		pat_enrollstudy_ops = result_pat["EnrollStudy_Ops"]
		# no_action
		# update pat_id, src_sam_id, UUID where src_id,src_pat_id
		# insert src_id, src_pat_id, pat_id, src_sam_id, UUID

		sam_enrollstudy_ops = result_sam["EnrollStudy_Ops"] 
		# update src_pat_id, pat_id, UUID where src_id, src_sam_id
		# insert src_id, src_pat_id, pat_id, src_sam_id, UUID

		enroll_study_data = {'Study_ID': sample_data['Study_ID'], \
							'Patient_ID': pat_id, \
							'Within_Study_Patient_ID': sample_data['Within_Study_Patient_ID'] , \
							'Sample_UUID': sample_data["UUID"], \
							'Within_Study_Sample_ID': sample_data['Within_Study_Sample_ID']}

		is_new_enroll_study = False
		last_enroll_study_id = None
		if((pat_enrollstudy_ops == "insert") and (sam_enrollstudy_ops == "insert")):
			is_new_enroll_study= True
			last_enroll_study_id = addEnrollStudyInDB(enroll_study_data)
		elif(sam_enrollstudy_ops == "update"):
			is_success = updateEnrollStudyInDB(enroll_study_data, by_variable = "Within_Study_Sample_ID")
			if(not is_success):
				isError = True
				errorMsg = errorMsg.join("Please contact developer. Fail to update new enroll study into database based on your input.")
				break
		else:
			#(sam_erollstudy_ops = "insert" and (pat_erollstudy_ops == "update" or pat_erollstudy_ops == "no_action")):
			is_success =  updateEnrollStudyInDB(enroll_study_data, by_variable = "Within_Study_Patient_ID")
			if(not is_success):
				isError = True
				errorMsg = errorMsg.join("Please contact developer. Fail to update new enroll study into database based on your input.")
				break


		addSucessResultIntoRemoteRDB(job_id = job_id, sample_uuid = sample_data["UUID"] , is_new_patient = is_new_patient, patient_id = pat_id, is_new_enroll_study = is_new_enroll_study, enroll_study_id = last_enroll_study_id)

		rowNum += 1

			
	# 10/15/2018 - Currently results are stored in RemoteR database (for Python code - handle XLSX). 
	# 			   previously stored in SMS database (for previous version PHP code - handle CSV).	 
	if(isError):
		addErrorResultIntoRemoteRDB(job_id = job_id, msg = errorMsg)	

	if(warnMsg != ""):
		addWarnResultIntoRemoteRDB(job_id = job_id, msg = warnMsg, existed_sam_uuids = existedUUIDs)

	os.remove(tmp_xlsx_filename)
	#10/16/2018 - generate an empty success.txt output file and let perl script to detect it
	# p.s. Which one is faster? 1) Let perl script to detect output file OR 2) to query remoter db for result multiple times until ~ 10 sec?	
		
	write_file("success","tmp_sam.txt")

if __name__ == '__main__':
    main()
